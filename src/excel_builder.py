"""
excel_builder.py
Builds a professional multi-sheet CRM Excel workbook with formatting and formulas.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA_DIRECTORY   = Path(__file__).parent.parent / "data"
OUTPUTS_DIRECTORY = Path(__file__).parent.parent

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_NAVY_COLOR  = "0D1117"
MID_NAVY_COLOR   = "161B22"
ACCENT_BLUE      = "1F6FEB"
SUCCESS_GREEN    = "2EA043"
DANGER_RED       = "DA3633"
WARNING_AMBER    = "D29922"
WHITE_COLOR      = "F0F6FC"
LIGHT_GREY_COLOR = "C9D1D9"
BORDER_COLOR     = "30363D"

HEADER_FILL      = PatternFill("solid", fgColor=ACCENT_BLUE)
ALTERNATE_FILL   = PatternFill("solid", fgColor="0E1724")
DARK_FILL        = PatternFill("solid", fgColor=DARK_NAVY_COLOR)
MIDTONE_FILL     = PatternFill("solid", fgColor=MID_NAVY_COLOR)

HEADER_FONT      = Font(name="Arial", bold=True, color=WHITE_COLOR, size=10)
BODY_FONT        = Font(name="Arial", color=LIGHT_GREY_COLOR, size=10)
TITLE_FONT       = Font(name="Arial", bold=True, color=WHITE_COLOR, size=14)
LABEL_FONT       = Font(name="Arial", bold=True, color=LIGHT_GREY_COLOR, size=10)

THIN_BORDER_SIDE = Side(style="thin", color=BORDER_COLOR)
CELL_BORDER      = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE,
)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGNMENT  = Alignment(horizontal="right",  vertical="center")
LEFT_ALIGNMENT   = Alignment(horizontal="left",   vertical="center")

PIPELINE_STAGE_ORDER = ["Lead", "Qualified", "Demo", "Proposal", "Negotiation"]
STAGE_WIN_PROBABILITY_MAP = {
    "Lead":        0.05,
    "Qualified":   0.15,
    "Demo":        0.30,
    "Proposal":    0.50,
    "Negotiation": 0.75,
}


# ── Helper functions ──────────────────────────────────────────────────────────

def write_header_cell(worksheet, row_number, column_number, cell_value, column_width=None):
    """Write a styled header cell with blue background and bold white text."""
    header_cell               = worksheet.cell(row=row_number, column=column_number, value=cell_value)
    header_cell.font          = HEADER_FONT
    header_cell.fill          = HEADER_FILL
    header_cell.border        = CELL_BORDER
    header_cell.alignment     = CENTER_ALIGNMENT
    if column_width:
        worksheet.column_dimensions[get_column_letter(column_number)].width = column_width
    return header_cell


def write_data_cell(worksheet, row_number, column_number, cell_value=None,
                    number_format=None, is_bold=False, font_color=LIGHT_GREY_COLOR,
                    background_fill=None, text_alignment=LEFT_ALIGNMENT):
    """Write a styled data cell with dark background and optional formatting."""
    data_cell               = worksheet.cell(row=row_number, column=column_number, value=cell_value)
    data_cell.font          = Font(name="Arial", color=font_color, bold=is_bold, size=10)
    data_cell.fill          = background_fill or MIDTONE_FILL
    data_cell.border        = CELL_BORDER
    data_cell.alignment     = text_alignment
    if number_format:
        data_cell.number_format = number_format
    return data_cell


def write_section_title(worksheet, row_number, column_number, title_text, column_span=8):
    """Write a full-width section title that merges across multiple columns."""
    title_cell            = worksheet.cell(row=row_number, column=column_number, value=title_text)
    title_cell.font       = TITLE_FONT
    title_cell.fill       = PatternFill("solid", fgColor="0A0E15")
    title_cell.alignment  = LEFT_ALIGNMENT
    worksheet.merge_cells(
        start_row=row_number, start_column=column_number,
        end_row=row_number,   end_column=column_number + column_span - 1,
    )
    worksheet.row_dimensions[row_number].height = 28


# ── Sheet 1: Executive Dashboard ─────────────────────────────────────────────

def build_dashboard_sheet(workbook: Workbook, deals_dataframe: pd.DataFrame,
                           forecast_dataframe: pd.DataFrame, leaderboard_dataframe: pd.DataFrame):
    dashboard_sheet = workbook.active
    dashboard_sheet.title = "📊 Dashboard"
    dashboard_sheet.sheet_view.showGridLines = False
    dashboard_sheet.sheet_view.zoomScale = 90

    # Title banner spanning A1:H2
    dashboard_sheet.merge_cells("A1:H2")
    banner_cell           = dashboard_sheet["A1"]
    banner_cell.value     = "CRM SALES PIPELINE — EXECUTIVE DASHBOARD"
    banner_cell.font      = Font(name="Arial", bold=True, color=WHITE_COLOR, size=16)
    banner_cell.fill      = PatternFill("solid", fgColor=ACCENT_BLUE)
    banner_cell.alignment = CENTER_ALIGNMENT
    dashboard_sheet.row_dimensions[1].height = 35
    dashboard_sheet.row_dimensions[2].height = 10

    # KPI summary cards
    won_deals   = deals_dataframe[deals_dataframe["stage"] == "Closed Won"]
    open_deals  = deals_dataframe[~deals_dataframe["stage"].isin(["Closed Won", "Closed Lost"])]
    closed_deals = deals_dataframe[deals_dataframe["stage"].isin(["Closed Won", "Closed Lost"])]
    overall_win_rate = len(won_deals) / len(closed_deals) if len(closed_deals) > 0 else 0

    kpi_card_definitions = [
        ("Total Pipeline",  f"${open_deals['acv'].sum() / 1e6:.1f}M",  ACCENT_BLUE),
        ("Closed Won ACV",  f"${won_deals['acv'].sum() / 1e6:.1f}M",   SUCCESS_GREEN),
        ("Avg Deal Size",   f"${won_deals['acv'].mean():,.0f}",         WARNING_AMBER),
        ("Win Rate",        f"{overall_win_rate:.1%}",                  ACCENT_BLUE),
        ("Open Deals",      str(len(open_deals)),                       WARNING_AMBER),
        ("Top Rep",         leaderboard_dataframe.index[0],             SUCCESS_GREEN),
    ]

    for card_index, (card_label, card_value, value_color) in enumerate(kpi_card_definitions):
        column_number = card_index + 1
        dashboard_sheet.column_dimensions[get_column_letter(column_number)].width = 18
        dashboard_sheet.row_dimensions[4].height = 16
        dashboard_sheet.row_dimensions[5].height = 30
        dashboard_sheet.row_dimensions[6].height = 16

        label_cell            = dashboard_sheet.cell(row=4, column=column_number, value=card_label)
        label_cell.font       = Font(name="Arial", color=LIGHT_GREY_COLOR, size=9, bold=True)
        label_cell.fill       = PatternFill("solid", fgColor="0A0E15")
        label_cell.alignment  = CENTER_ALIGNMENT

        value_cell            = dashboard_sheet.cell(row=5, column=column_number, value=card_value)
        value_cell.font       = Font(name="Arial", color=value_color, size=14, bold=True)
        value_cell.fill       = MIDTONE_FILL
        value_cell.alignment  = CENTER_ALIGNMENT
        value_cell.border     = CELL_BORDER

    # Pipeline summary table
    write_section_title(dashboard_sheet, 8, 1, "  📌  Pipeline Summary by Stage", 5)
    header_labels  = ["Stage", "Deals", "Total ACV", "Weighted ACV", "Win Prob"]
    column_widths  = [14, 8, 16, 16, 10]
    for column_index, (header_label, column_width) in enumerate(zip(header_labels, column_widths), 1):
        write_header_cell(dashboard_sheet, 9, column_index, header_label, column_width)

    for stage_index, stage_name in enumerate(PIPELINE_STAGE_ORDER):
        deals_in_stage  = deals_dataframe[deals_dataframe["stage"] == stage_name]
        data_row_number = 10 + stage_index
        row_fill        = ALTERNATE_FILL if stage_index % 2 else MIDTONE_FILL
        win_probability = STAGE_WIN_PROBABILITY_MAP[stage_name]
        prob_color      = SUCCESS_GREEN if win_probability >= 0.5 else WARNING_AMBER

        write_data_cell(dashboard_sheet, data_row_number, 1, stage_name,
                        is_bold=True, font_color=WHITE_COLOR, background_fill=row_fill)
        write_data_cell(dashboard_sheet, data_row_number, 2, len(deals_in_stage),
                        background_fill=row_fill, text_alignment=CENTER_ALIGNMENT)
        write_data_cell(dashboard_sheet, data_row_number, 3, deals_in_stage["acv"].sum(),
                        number_format='$#,##0', background_fill=row_fill,
                        text_alignment=RIGHT_ALIGNMENT)
        write_data_cell(dashboard_sheet, data_row_number, 4,
                        f'=C{data_row_number}*E{data_row_number}',
                        number_format='$#,##0', background_fill=row_fill,
                        text_alignment=RIGHT_ALIGNMENT)
        write_data_cell(dashboard_sheet, data_row_number, 5, win_probability,
                        number_format='0%', background_fill=row_fill,
                        text_alignment=CENTER_ALIGNMENT, font_color=prob_color)

    # Totals row
    total_row_number    = 10 + len(PIPELINE_STAGE_ORDER)
    total_row_fill      = PatternFill("solid", fgColor="0A0E15")
    write_data_cell(dashboard_sheet, total_row_number, 1, "TOTAL",
                    is_bold=True, font_color=WHITE_COLOR, background_fill=total_row_fill)
    write_data_cell(dashboard_sheet, total_row_number, 2,
                    f"=SUM(B10:B{total_row_number - 1})",
                    is_bold=True, text_alignment=CENTER_ALIGNMENT, background_fill=total_row_fill)
    write_data_cell(dashboard_sheet, total_row_number, 3,
                    f"=SUM(C10:C{total_row_number - 1})",
                    is_bold=True, number_format='$#,##0', text_alignment=RIGHT_ALIGNMENT,
                    background_fill=total_row_fill, font_color=WARNING_AMBER)
    write_data_cell(dashboard_sheet, total_row_number, 4,
                    f"=SUM(D10:D{total_row_number - 1})",
                    is_bold=True, number_format='$#,##0', text_alignment=RIGHT_ALIGNMENT,
                    background_fill=total_row_fill, font_color=SUCCESS_GREEN)
    dashboard_sheet.cell(total_row_number, 5).fill = total_row_fill

    # Forecast table
    write_section_title(dashboard_sheet, 8, 7, "  📈  Revenue Forecast 2025", 4)
    forecast_headers = ["Quarter", "Bear Case", "Base Case", "Bull Case"]
    forecast_widths  = [12, 14, 14, 14]
    for column_index, (forecast_header, forecast_width) in enumerate(
        zip(forecast_headers, forecast_widths), 1
    ):
        write_header_cell(dashboard_sheet, 9, 6 + column_index, forecast_header, forecast_width)

    for forecast_index, forecast_row in forecast_dataframe.iterrows():
        data_row_number = 10 + forecast_index
        row_fill        = ALTERNATE_FILL if forecast_index % 2 else MIDTONE_FILL
        write_data_cell(dashboard_sheet, data_row_number, 7,  forecast_row["quarter"],
                        is_bold=True, font_color=WHITE_COLOR, background_fill=row_fill)
        write_data_cell(dashboard_sheet, data_row_number, 8,  forecast_row["bear_case"],
                        number_format='$#,##0', text_alignment=RIGHT_ALIGNMENT,
                        background_fill=row_fill, font_color=DANGER_RED)
        write_data_cell(dashboard_sheet, data_row_number, 9,  forecast_row["base_case"],
                        number_format='$#,##0', text_alignment=RIGHT_ALIGNMENT,
                        background_fill=row_fill, font_color=WARNING_AMBER)
        write_data_cell(dashboard_sheet, data_row_number, 10, forecast_row["bull_case"],
                        number_format='$#,##0', text_alignment=RIGHT_ALIGNMENT,
                        background_fill=row_fill, font_color=SUCCESS_GREEN)


# ── Sheet 2: Deal Data ────────────────────────────────────────────────────────

def build_deal_data_sheet(workbook: Workbook, deals_dataframe: pd.DataFrame):
    deal_sheet = workbook.create_sheet("📋 Deal Data")
    deal_sheet.sheet_view.showGridLines = False
    deal_sheet.freeze_panes = "A2"

    column_names = [
        "deal_id", "deal_name", "rep_name", "rep_segment", "region", "industry",
        "product", "lead_source", "acv", "stage", "created_date", "close_date",
        "days_in_pipeline", "deal_score", "forecast_category", "competitors",
    ]
    column_widths = [10, 22, 16, 12, 14, 14, 18, 13, 12, 14, 13, 13, 16, 11, 16, 13]
    currency_and_number_formats = {8: '$#,##0', 13: '0.0', 14: '0.0'}

    for column_index, (column_name, column_width) in enumerate(
        zip(column_names, column_widths), 1
    ):
        write_header_cell(deal_sheet, 1, column_index,
                          column_name.replace("_", " ").title(), column_width)

    for row_index, (_, deal_row) in enumerate(deals_dataframe[column_names].iterrows(), 2):
        row_fill = ALTERNATE_FILL if row_index % 2 == 0 else MIDTONE_FILL
        for column_index, cell_value in enumerate(deal_row.values, 1):
            number_format = currency_and_number_formats.get(column_index)
            font_color    = WHITE_COLOR if column_index in (1, 2, 10) else LIGHT_GREY_COLOR
            write_data_cell(deal_sheet, row_index, column_index, cell_value,
                            background_fill=row_fill, number_format=number_format,
                            font_color=font_color)

    deal_sheet.auto_filter.ref = f"A1:{get_column_letter(len(column_names))}1"


# ── Sheet 3: Rep Performance ──────────────────────────────────────────────────

def build_rep_performance_sheet(workbook: Workbook, leaderboard_dataframe: pd.DataFrame,
                                 targets_dataframe: pd.DataFrame):
    rep_sheet = workbook.create_sheet("🏆 Rep Performance")
    rep_sheet.sheet_view.showGridLines = False

    write_section_title(rep_sheet, 1, 1, "  🏆  Sales Rep Leaderboard", 7)

    leaderboard_headers = ["Rep Name", "Won Deals", "Won ACV", "Avg Deal ($)",
                           "Win Rate", "Annual Quota", "Attainment"]
    leaderboard_widths  = [18, 10, 14, 14, 10, 14, 12]
    for column_index, (header_label, column_width) in enumerate(
        zip(leaderboard_headers, leaderboard_widths), 1
    ):
        write_header_cell(rep_sheet, 2, column_index, header_label, column_width)

    medal_icons = ["🥇", "🥈", "🥉"]
    for row_index, (rep_name, rep_row) in enumerate(leaderboard_dataframe.iterrows(), 3):
        row_fill         = ALTERNATE_FILL if row_index % 2 else MIDTONE_FILL
        attainment_value = rep_row["quota_attainment"]
        attainment_color = (
            SUCCESS_GREEN if attainment_value >= 1.0
            else WARNING_AMBER if attainment_value >= 0.75
            else DANGER_RED
        )
        medal_icon = medal_icons[row_index - 3] if row_index <= 5 else ""

        write_data_cell(rep_sheet, row_index, 1, f"{medal_icon} {rep_name}",
                        background_fill=row_fill, is_bold=row_index <= 5,
                        font_color=WHITE_COLOR)
        write_data_cell(rep_sheet, row_index, 2, int(rep_row["won_deal_count"]),
                        background_fill=row_fill, text_alignment=CENTER_ALIGNMENT)
        write_data_cell(rep_sheet, row_index, 3, rep_row["total_won_acv"],
                        background_fill=row_fill, number_format='$#,##0',
                        text_alignment=RIGHT_ALIGNMENT, font_color=SUCCESS_GREEN)
        write_data_cell(rep_sheet, row_index, 4, rep_row["average_deal_size"],
                        background_fill=row_fill, number_format='$#,##0',
                        text_alignment=RIGHT_ALIGNMENT)
        write_data_cell(rep_sheet, row_index, 5, rep_row["win_rate"],
                        background_fill=row_fill, number_format='0.0%',
                        text_alignment=CENTER_ALIGNMENT)
        write_data_cell(rep_sheet, row_index, 6, rep_row.get("annual_quota", 0),
                        background_fill=row_fill, number_format='$#,##0',
                        text_alignment=RIGHT_ALIGNMENT)
        write_data_cell(rep_sheet, row_index, 7, attainment_value,
                        background_fill=row_fill, number_format='0.0%',
                        text_alignment=CENTER_ALIGNMENT, font_color=attainment_color)

    # Monthly quota vs actual sub-table
    monthly_summary = targets_dataframe.groupby("month")[["quota", "actual"]].sum().reset_index()
    monthly_summary["month"] = pd.to_datetime(monthly_summary["month"]).dt.strftime("%b %Y")
    monthly_table_start_row  = len(leaderboard_dataframe) + 5

    write_section_title(rep_sheet, monthly_table_start_row, 1,
                        "  📅  Monthly Quota vs Actual (All Reps)", 4)
    monthly_headers = ["Month", "Total Quota", "Total Actual", "Attainment"]
    for column_index, header_label in enumerate(monthly_headers, 1):
        write_header_cell(rep_sheet, monthly_table_start_row + 1, column_index, header_label)

    for row_index, monthly_row in monthly_summary.iterrows():
        data_row_number  = monthly_table_start_row + 2 + row_index
        monthly_fill     = ALTERNATE_FILL if row_index % 2 else MIDTONE_FILL
        attainment_rate  = (
            monthly_row["actual"] / monthly_row["quota"]
            if monthly_row["quota"] > 0 else 0
        )
        attainment_color = (
            SUCCESS_GREEN if attainment_rate >= 1.0
            else WARNING_AMBER if attainment_rate >= 0.8
            else DANGER_RED
        )

        write_data_cell(rep_sheet, data_row_number, 1, monthly_row["month"],
                        background_fill=monthly_fill, is_bold=True, font_color=WHITE_COLOR)
        write_data_cell(rep_sheet, data_row_number, 2, monthly_row["quota"],
                        background_fill=monthly_fill, number_format='$#,##0',
                        text_alignment=RIGHT_ALIGNMENT)
        write_data_cell(rep_sheet, data_row_number, 3, monthly_row["actual"],
                        background_fill=monthly_fill, number_format='$#,##0',
                        text_alignment=RIGHT_ALIGNMENT)
        write_data_cell(rep_sheet, data_row_number, 4, f"=C{data_row_number}/B{data_row_number}",
                        background_fill=monthly_fill, number_format='0.0%',
                        text_alignment=CENTER_ALIGNMENT, font_color=attainment_color)


# ── Sheet 4: Tableau Export ───────────────────────────────────────────────────

def build_tableau_export_sheet(workbook: Workbook, deals_dataframe: pd.DataFrame,
                                activities_dataframe: pd.DataFrame):
    """
    Flat, clean table optimised for Tableau connection.
    Includes all dimensions and enriched calculated fields.
    """
    tableau_sheet = workbook.create_sheet("📤 Tableau Export")
    tableau_sheet.sheet_view.showGridLines = False

    enriched_deals = deals_dataframe.copy()
    enriched_deals["is_won"]          = (enriched_deals["stage"] == "Closed Won").astype(int)
    enriched_deals["is_closed"]       = enriched_deals["stage"].isin(
        ["Closed Won", "Closed Lost"]
    ).astype(int)
    enriched_deals["win_probability"] = enriched_deals["stage"].map({
        "Lead": 0.05, "Qualified": 0.15, "Demo": 0.30, "Proposal": 0.50,
        "Negotiation": 0.75, "Closed Won": 1.0, "Closed Lost": 0.0,
    })
    enriched_deals["weighted_acv"]    = (
        enriched_deals["acv"] * enriched_deals["win_probability"]
    ).round(0)
    enriched_deals["created_year"]    = pd.to_datetime(enriched_deals["created_date"]).dt.year
    enriched_deals["created_quarter"] = (
        pd.to_datetime(enriched_deals["created_date"]).dt.to_period("Q").astype(str)
    )
    enriched_deals["created_month"]   = (
        pd.to_datetime(enriched_deals["created_date"]).dt.to_period("M").astype(str)
    )

    activity_count_per_deal = (
        activities_dataframe.groupby("deal_id")["activity_id"]
        .count()
        .rename("activity_count")
    )
    enriched_deals = enriched_deals.join(activity_count_per_deal, on="deal_id")

    export_column_names = [
        "deal_id", "deal_name", "rep_name", "rep_segment", "region", "industry",
        "product", "lead_source", "acv", "weighted_acv", "win_probability", "stage",
        "is_won", "is_closed", "created_date", "close_date", "days_in_pipeline",
        "deal_score", "num_touches", "activity_count", "competitors",
        "forecast_category", "created_year", "created_quarter", "created_month",
    ]
    export_ready_dataframe = enriched_deals[export_column_names].fillna("")

    for column_index, column_name in enumerate(export_column_names, 1):
        write_header_cell(
            tableau_sheet, 1, column_index,
            column_name.replace("_", " ").title(),
            max(len(column_name) + 2, 12),
        )

    for row_index, (_, export_row) in enumerate(export_ready_dataframe.iterrows(), 2):
        row_fill = ALTERNATE_FILL if row_index % 2 == 0 else MIDTONE_FILL
        for column_index, cell_value in enumerate(export_row.values, 1):
            write_data_cell(tableau_sheet, row_index, column_index,
                            cell_value, background_fill=row_fill)

    tableau_sheet.auto_filter.ref = f"A1:{get_column_letter(len(export_column_names))}1"
    tableau_sheet.freeze_panes = "A2"


# ── Main builder ──────────────────────────────────────────────────────────────

def build_excel_workbook(deals_dataframe: pd.DataFrame, forecast_dataframe: pd.DataFrame,
                          leaderboard_dataframe: pd.DataFrame, targets_dataframe: pd.DataFrame,
                          activities_dataframe: pd.DataFrame) -> Path:
    crm_workbook = Workbook()

    build_dashboard_sheet(crm_workbook, deals_dataframe, forecast_dataframe, leaderboard_dataframe)
    build_deal_data_sheet(crm_workbook, deals_dataframe)
    build_rep_performance_sheet(crm_workbook, leaderboard_dataframe, targets_dataframe)
    build_tableau_export_sheet(crm_workbook, deals_dataframe, activities_dataframe)

    output_file_path = DATA_DIRECTORY / "CRM_Sales_Pipeline.xlsx"
    crm_workbook.save(output_file_path)
    print(f"✅ Excel workbook saved  →  {output_file_path}")
    return output_file_path